#This script is being used to extract the Application data type mapping from mapping set
#For Every implementation data types there is a mapping for application datatype

#input: Excel Sheet Where the first and second coulmn contaion the calibration parameters
#output : Fill Coulumn 5 and 6 with the Application datatypes
#output : Fill Coulumn 7 and 8 with Implementation datatypes

#import libraries being used
import os
import openpyxl

#Working
#SW version 2.5

#add folder that contain .arxml file 
arxml_DataTypes =  "D:\DataTypes.arxml"

#add path that contain .xlsx file which have data type
excel_DataTypes =  "D:/table_info_Data_Stage_B_PATH_3.xlsx"

# Create temp file to copy the arxmls, and edit in the new file 
#data_index = arxml_DataTypes.find("\\")
#new_arxml_DataTypes = arxml_DataTypes[:(data_index)] +"/new_"+arxml_DataTypes[(data_index+1):]

# workbook object is created 
wb_obj = openpyxl.load_workbook(excel_DataTypes)

#create Sheet object
sheet_obj = wb_obj.get_sheet_by_name('1D_Tables')

#get max numbers of rows
maxmium_row = sheet_obj.max_row

#declare a list to save lines number in it , initalize by zeros
DataTypes_Found_At_Lines_First_Column = [0] * maxmium_row
DataTypes_Found_At_Lines_Second_Column = [0] * maxmium_row

#The Implementation Datatype we are searching for start and end lines
Line_Start = '<IMPLEMENTATION-DATA-TYPE-REF'
Line_End = '</IMPLEMENTATION-DATA-TYPE-REF>'

DATA_TYPE_MAP_tag = 0
Current_line_where_is_tag = 0

#Step 1
#For Loop For The two Columns
for j in range(1,3):
    #for loop for the max number of row in the sheet , serach one by one
    for i in range(1, maxmium_row + 1 ):
        with open(arxml_DataTypes,'r') as inFile:
            
            #Get Cell object Data
            cell_obj = sheet_obj.cell(row = i, column = j)
            #search in the whole file , Sequencal search Slow $need to be updated to binary search for example
            for num_line, line_content in enumerate(inFile, 1):

                #get calibration Parameters from excel sheet
                DataType = str(cell_obj.value)
                
                #search for Data Type Map Tag
                if line_content.find('<DATA-TYPE-MAP>') != -1 :
                    #Set Data map tag flag to 1
                    DATA_TYPE_MAP_tag = 1
                    Current_line_where_is_tag = num_line
                #if three lines passed and couldn't find data remove flag
                elif(num_line > Current_line_where_is_tag + 3 ) :
                    DATA_TYPE_MAP_tag = 0
                # check for required data mapping
                elif line_content.find(Line_Start) != -1 and line_content.find('Idt_') and line_content.find(DataType) != -1 and line_content.find(Line_End) != -1 and DATA_TYPE_MAP_tag == 1:
                    #found the Implementation data type But flag = 1
                    Found_Datatype_Flag = 1
                    DATA_TYPE_MAP_tag = 0
                    if(j == 1):
                        #Save Line numbers of the found Elements in first ROW
                        DataTypes_Found_At_Lines_First_Column.insert (i,num_line)
                        #search for the place of implementation in the line itself
                        X = line_content.find('Implementation_Types/')
                        Y = line_content.find('</IMPLEMENTATION-DATA-TYPE-REF>')
                        #assign value to it
                        Implementation_Types = line_content[X + 21 : Y]
                        #write the Implemenation datatype to the correct column in the excel sheet
                        sheet_obj.cell(row = i, column = 7).value = Implementation_Types
                        break
                    if(j == 2):
                        #Save Line numbers of the found Elements in first ROW
                        DataTypes_Found_At_Lines_Second_Column.insert (i,num_line)
                        #search for the place of implementation in the line itself
                        X = line_content.find('Implementation_Types/')
                        Y = line_content.find('</IMPLEMENTATION-DATA-TYPE-REF>')
                        #assign value to it
                        Implementation_Types = line_content[X + 21 : Y]
                        #write the Implemenation datatype to the correct column in the excel sheet
                        sheet_obj.cell(row = i, column = 8).value = Implementation_Types
                        break

                #couldn't found the calibartion implementation data set    
                else:
                    Found_Datatype_Flag = 0
            
            if(Found_Datatype_Flag != 1):
                print("Couldn't find : "+ str(cell_obj.value))
#Print First Column lines number
print(DataTypes_Found_At_Lines_First_Column)
#print Second Column Lines number
print(DataTypes_Found_At_Lines_Second_Column)

#Complete Step 1
#Get Application Data Type
for j in range(1,3):
    #For Loop for the number of rows
    for i in range(1, maxmium_row + 1 ):
        with open(arxml_DataTypes,'r') as inFile:
            #Get Line numbers - 1
            for num_line, line_content in enumerate(inFile, 1):
                #serach in the first column
                if j == 1:
                    #we know the lines then get the data
                    if num_line == DataTypes_Found_At_Lines_First_Column[i] - 1 and DataTypes_Found_At_Lines_First_Column[i] != 0 :
                        #Subtaract Application datatype from Line content
                        X = line_content.find('Application_Types/')
                        Y = line_content.find('</APPLICATION-DATA-TYPE-REF>')
                        #Get AppDataType for line
                        App_DataType = line_content[X + 18 : Y]
                        #print Application datatype
                        #print(App_DataType)
                        #Save to the Excel Sheet Coulmn 5 for X axis
                        sheet_obj.cell(row = i, column = 5).value = App_DataType
                        wb_obj.save(excel_DataTypes)
                #Search in the second column
                if j == 2:
                    #we know the lines so we get the data
                    if num_line == DataTypes_Found_At_Lines_Second_Column[i] - 1 and DataTypes_Found_At_Lines_Second_Column[i] != 0 :
                        #Subtaract Application datatype from Line content
                        X = line_content.find('Application_Types/')
                        Y = line_content.find('</APPLICATION-DATA-TYPE-REF>')
                        #Get AppDataType for line
                        App_DataType = line_content[X + 18 : Y]
                        #print Application datatype
                        #print(App_DataType)
                        #Save to the Excel Sheet Coulmn 5 for X axis
                        sheet_obj.cell(row = i, column = 6).value = App_DataType
                        wb_obj.save(excel_DataTypes)
#print message that we finished getting Application datatypes
print("Finished Getting Application Data types for every implemention Type ")
