#This script is being used to Rename Implementation data type to Application data Type
#input Excel Sheet must be Like :
#First Three Coulmns contains X , Y , Z Cals
#Fourth column contains SWC Name , Must match the file path name

#output :
#All Done Rows Will be Mapped to Done 
#Files will be changed

#import libraries being used
import os
import openpyxl
import glob
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors

#add folder that contain .arxml files
SWC_Arxmls = "D:/Avelabs/Ford_Dat2.1/aptiv_sw/autosar_cfg/davinci/Config/Developer/ComponentTypes/*.arxml"

#Get All Paths for Arxmls in the folder
SWC_Arxmls_Path = glob.glob(SWC_Arxmls)

#add path that contain .xlsx file which have data type
excel_DataTypes =  "D:/table_info_Data_Stage_B_PATH_3.xlsx"

#Start Row
Start_Row = 2

#Line Start
Line_Start = '<TYPE-TREF DEST="IMPLEMENTATION-DATA-TYPE">'
#Middle Line
Middle_Line = 'Cal_Datatype'
#Line End
Line_End = '</TYPE-TREF>'

#Global variable to hold column number
Column = 0
#Global Variable to hold row number
row = 0

def main():
    #Load WorkBook
    Workbook_Object = Load_Workbook(excel_DataTypes)
    # workbook object is created , create Sheet object
    Sheet_object = Load_Sheet(Workbook_Object , '1D_Tables')
    #get max numbers of rows
    maxmium_row = Get_Max_Row(Sheet_object) + 1
    #Todo Change to 4 in case of 3D Cals
    for Column in range(1 , 3 ):
        for row in range(Start_Row , maxmium_row ) :
            #get Cal Parameters from excel sheet
            Cal_Name = Get_Value_Of_Cell(Sheet_object , row , Column )
            #get Implementation Parameters from excel sheet
            Imp_DataType = Get_Value_Of_Cell(Sheet_object , row , Column + 7 )
            #get application Type
            Application_Type = Get_Value_Of_Cell(Sheet_object , row , Column + 4 )
            #get SWC Name
            SWC_Name = Get_Value_Of_Cell(Sheet_object , row ,  4 )
            #Loop For all Arxmls
            for SWC_arxml in SWC_Arxmls_Path:
                #Search For SWC Needed
                if SWC_arxml.find(SWC_Name) != -1:
                    # Create temp file to copy the arxmls, and edit in the new file
                    data_index = SWC_arxml.find("\\")
                    new_SWC_arxml = SWC_arxml[:(data_index)] +"/new_"+SWC_arxml[(data_index+1):]
                    #open the new file
                    new_file = open(new_SWC_arxml,'w')
                    #open the old file
                    with open(SWC_arxml,'r') as inFile:
                        #Search in the whole file
                        for num_line , line_content in enumerate(inFile, 1):
                            #if found the the Cal Name
                            if line_content.find('<SHORT-NAME>'+Cal_Name+'</SHORT-NAME>') != -1:
                                Cal_Name_Found_Flag  = 1
                            #Replace
                            if (line_content.find(Imp_DataType) != -1) and (Cal_Name_Found_Flag == 1) :
                                if (Application_Type != 'None'):
                                    #rename implementation to Application
                                    line_content = line_content.replace(Imp_DataType, Application_Type)
                                    #Change other headers
                                    line_content = line_content.replace('"IMPLEMENTATION-DATA-TYPE">' , '"APPLICATION-PRIMITIVE-DATA-TYPE">')
                                    Component_Start_in_line = line_content.find('ComponentType/')
                                    Component_End_in_line = line_content.find('Cal_Datatype/')
                                    line_content = line_content.replace(line_content[Component_Start_in_line : Component_End_in_line + 13 ],'Data_Type/Application_Types/')
                                    #Write Done To Excel Sheet
                                    Write_Value_To_Cell(Sheet_object , row , Column + 10 , 'Done' )
                                    #Color Cell with Green
                                    Color_Cell_Green(Sheet_object , row , Column + 10)
                                    #Save WorkObject "Excel Sheet After Edits"
                                    Workbook_Object.save(excel_DataTypes)
                                    #Print Line in Console
                                    print(line_content)
                            #Found Tag Closing
                            if line_content.find('</PARAMETER-DATA-PROTOTYPE>') != -1:
                                Cal_Name_Found_Flag = 0
                            #Write Line
                            new_file.write(line_content)
                    #Close the File after Edits
                    new_file.close()
                    #Copt New File to the old file every Line Change
                    Copy_File_Content(new_SWC_arxml , SWC_arxml )
                    # remove the orignal file to replace the new one with it   
                    os.remove(SWC_arxml)
                    # rename the new file to have the same name of the orignal one 
                    os.rename(new_SWC_arxml,SWC_arxml)
    #Done
    print("Renaming Completed Successfully")


#####Local Functions

#Function To Load WorkBook and Return Work Book Object
def Load_Workbook(Input_Path):
    Workbook_Obj = openpyxl.load_workbook(Input_Path)
    return Workbook_Obj

#Function that takes input path , sheet name , and create a object from the sheet
def Load_Sheet( Workbook_Object , Sheet_Name):
    Sheet_Object = Workbook_Object.get_sheet_by_name(Sheet_Name)
    return Sheet_Object

#function that takes a sheet object and return max row
def Get_Max_Row(sheet_obj):
    Max = sheet_obj.max_row
    return Max

#Function that takes sheet object , ROW , Coulumn
#Return value of a cell
def Get_Value_Of_Cell(sheet_obj , row , column):
    Cell_objj = sheet_obj.cell(row = row , column = column)
    Cell_Value = str(Cell_objj.value)
    return Cell_Value

#write value to any cell in the excel sheet
def Write_Value_To_Cell(sheet_obj , row , column , value):
    sheet_obj.cell(row = row , column = column).value = value

#Function that copies All contenet of a file and copies to an other file
def Copy_File_Content(src , dest):
    with open(src,'r') as F1:
        with open(dest,'w') as F2:
            for Line in F1:
                F2.write(Line)
    F1.close()
    F2.close()

#Color Cell With Red
def Color_Cell_Red( sheet_object , Row , Column ):
    redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    Cell = sheet_object.cell(row = Row , column = Column)
    Cell.fill = redFill

#Color Cell With Green
def Color_Cell_Green( sheet_object , Row , Column ):
    Green_Fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    Cell = sheet_object.cell(row = Row , column = Column)
    Cell.fill = Green_Fill
#Color Cell With Yellow
def Color_Cell_Yellow ( sheet_object , Row , Column ):
    Yellow_Fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    Cell = sheet_object.cell(row = Row , column = Column)
    Cell.fill = Yellow_Fill


if __name__ == '__main__':
    main()