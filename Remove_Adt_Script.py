#This Script is used to Remove Adt_& _T from datatypes Arxml File for calbibration parameters
#It removes from two places <DATA-TYPE-MAP> & <APPLICATION-PRIMITIVE-DATA-TYPE
#input : Excel sheet containing calibration names
#output : Application datatypes is same name as calibration names

import os
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors

#add folder that contain .arxml file 
arxml_DataTypes =  "D:/Workspaces/DAT2.1_Workspaces/RTE_Workspace/aptiv_sw/autosar_cfg/davinci/Config/Developer/DataTypes.arxml"
new_Datatypes_arxml = "D:/Workspaces/DAT2.1_Workspaces/RTE_Workspace/aptiv_sw/autosar_cfg/davinci/Config/Developer/new_DataTypes.arxml"
#add path that contain .xlsx file which have data type
excel_DataTypes =  "D:/table_info_Data_Stage_B_PATH_3.xlsx"
# workbook object is created 
wb_obj = openpyxl.load_workbook(excel_DataTypes)
#create Sheet object
Sheet_object = wb_obj.get_sheet_by_name('1D_Tables')
#get max numbers of rows
maxmium_row = Sheet_object.max_row

#Column_Max_Number = 3
Column_Max_Number = 3
#Start_Row
Start_Row = 1
#Flag for DATA_TYPE_MAP_Line_Start_found
DATA_TYPE_MAP_Line_Start_found = 0


#The APPLICATION Datatype we are searching for start and end lines
APPLICATION_Line_Start = '<APPLICATION-PRIMITIVE-DATA-TYPE'
APPLICATION_Line_End = '</APPLICATION-PRIMITIVE-DATA-TYPE>'
#Data type map tags
DATA_TYPE_MAP_Line_Start = '<DATA-TYPE-MAP>'
DATA_TYPE_MAP_Line_End = '</DATA-TYPE-MAP>'

def main():
    for Column in range(1 , Column_Max_Number ):
        for row in range(Start_Row , maxmium_row ) :
            #open the new file
            new_file = open(new_Datatypes_arxml,'w')
            #get Cal Parameters from excel sheet
            Cal_Name = Get_Value_Of_Cell(Sheet_object , row , Column )
            with open(arxml_DataTypes,'r') as inFile:
                #Get Line numbers - 1
                for num_line, line_content in enumerate(inFile, 1):
                    #Search for opening of Data-type-map opening tag
                    if line_content.find(DATA_TYPE_MAP_Line_Start) != -1 :
                        #if found set a flage 
                        DATA_TYPE_MAP_Line_Start_found = 1
                    #if cal found
                    if((line_content.find('Adt_' + Cal_Name + '_T') != -1) and (DATA_TYPE_MAP_Line_Start_found == 1) and (line_content.find('<APPLICATION-DATA-TYPE-REF') != -1 )):
                        #Remove Adt_ and also _T From the Line
                        line_content = line_content.replace('Adt_' + Cal_Name + '_T', Cal_Name )
                        #Color the Cell 11 and 12
                        Write_Value_To_Cell(Sheet_object , row , Column + 11 , 'Deleted' )
                        Color_Cell_Green( Sheet_object , row , Column + 11 )
                        #Save WorkObject "Excel Sheet After Edits"
                        wb_obj.save(excel_DataTypes)
                        #print line
                        print(line_content)
                    #Remove Flage for closing flag
                    if line_content.find(DATA_TYPE_MAP_Line_End) != -1:
                        DATA_TYPE_MAP_Line_Start_found = 0
                    #Edit Name itself
                    if(line_content.find('<APPLICATION-PRIMITIVE-DATA-TYPE') != -1 ):
                        #Found APPLICATION-PRIMITIVE Type , Set a flag
                        APPLICATION_PRIMITIVE = 1
                    if ((line_content.find('<SHORT-NAME>' + 'Adt_' + Cal_Name + '_T' + '</SHORT-NAME>' ) != -1) and (APPLICATION_PRIMITIVE == 1)):
                        #Edit Name itself
                        #line_content = '<SHORT-NAME>' + Cal_Name + '</SHORT-NAME>\n'
                        line_content = line_content.replace('<SHORT-NAME>' + 'Adt_' + Cal_Name + '_T' + '</SHORT-NAME>', '<SHORT-NAME>' + Cal_Name + '</SHORT-NAME>')
                        #Color the Cell 13 and 14
                        Color_Cell_Red( Sheet_object , row , Column + 13 )
                        Write_Value_To_Cell(Sheet_object , row , Column + 13 , 'Deleted' )
                        #Save WorkObject "Excel Sheet After Edits"
                        wb_obj.save(excel_DataTypes)
                        #Print Line Content
                        print(line_content)
                    if line_content.find('Adt_' + Cal_Name + '_T</SHARED-AXIS-TYPE-REF>' ) != -1  :
                        line_content = line_content.replace('Adt_' + Cal_Name + '_T</SHARED-AXIS-TYPE-REF>', Cal_Name + '</SHARED-AXIS-TYPE-REF>')
                        #Color Axis Cell
                        Color_Cell_Red( Sheet_object , row , Column + 15 )
                        Write_Value_To_Cell(Sheet_object , row , Column + 15 , 'Axis' )
                        #Save WorkObject "Excel Sheet After Edits"
                        wb_obj.save(excel_DataTypes)
                        print(line_content)
                    #Closing Tag
                    if line_content.find('</APPLICATION-PRIMITIVE-DATA-TYPE>') != -1:
                        APPLICATION_PRIMITIVE = 0

                    #Write Line
                    new_file.write(line_content)
            #Close the File after Edits
            new_file.close()
            #Copt New File to the old file every Line Change
            Copy_File_Content(new_Datatypes_arxml , arxml_DataTypes )

#Function that takes sheet object , ROW , Coulumn
#Return value of a cell
def Get_Value_Of_Cell(sheet_obj , row , column):
    Cell_objj = sheet_obj.cell(row = row , column = column)
    Cell_Value = str(Cell_objj.value)
    return Cell_Value

#Function that copies All contenet of a file and copies to an other file
def Copy_File_Content(src , dest):
    with open(src,'r') as F1:
        with open(dest,'w') as F2:
            for Line in F1:
                F2.write(Line)
    F1.close()
    F2.close()

#Color Cell With Red
def Color_Cell_Red( sheet_object , Row , Column ):
    redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    Cell = sheet_object.cell(row = Row , column = Column)
    Cell.fill = redFill

#Color Cell With Green
def Color_Cell_Green( sheet_object , Row , Column ):
    Green_Fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    Cell = sheet_object.cell(row = Row , column = Column)
    Cell.fill = Green_Fill
#write value to any cell in the excel sheet
def Write_Value_To_Cell(sheet_obj , row , column , value):
    sheet_obj.cell(row = row , column = column).value = value

#Start App
if __name__ == '__main__':
    main()
#print message that we finished getting Application datatypes
print("Finished Remvoing Adt _T ")